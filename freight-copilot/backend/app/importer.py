"""Bulk freight-offer import from CSV / XLSX.

Two modes per row:
  * structured columns (origin_city, dest_city, weight_kg, ...), or
  * a free-text column (raw_text / text / opis / offer) parsed by the Intake
    agent — which also runs the prompt-injection safety scan.

All imported content is treated as untrusted data. Upload size and file-type
limits are enforced by the caller (see config + API route).
"""

from __future__ import annotations

import csv
import io

# Canonical field -> accepted header aliases (lowercased).
HEADER_ALIASES: dict[str, set[str]] = {
    "raw_text": {"raw_text", "text", "opis", "offer", "description", "tresc"},
    "source_reference": {"source_reference", "reference", "ref", "id", "nr"},
    "origin_city": {"origin_city", "origin", "from", "zaladunek", "loading", "od"},
    "origin_country": {"origin_country", "from_country", "kraj_zaladunku"},
    "dest_city": {"dest_city", "destination", "to", "rozladunek", "unloading", "do"},
    "dest_country": {"dest_country", "to_country", "kraj_rozladunku"},
    "weight_kg": {"weight_kg", "weight", "waga", "masa", "kg"},
    "vehicle_type": {"vehicle_type", "vehicle", "pojazd", "naczepa", "truck"},
    "adr_required": {"adr_required", "adr"},
    "customer_rate": {"customer_rate", "rate", "stawka", "price", "cena", "freight"},
    "currency": {"currency", "waluta", "cur"},
    "pickup_date": {"pickup_date", "date", "data", "loading_date", "termin"},
}

# Reverse lookup: alias -> canonical.
_ALIAS_TO_CANONICAL = {
    alias: canon for canon, aliases in HEADER_ALIASES.items() for alias in aliases
}

_TRUTHY = {"1", "true", "yes", "y", "tak", "adr", "x"}


class ImportError_(ValueError):
    """Raised for unparseable or unsupported import files."""


def parse_file(filename: str, content: bytes) -> list[dict]:
    """Parse CSV or XLSX bytes into a list of raw row dicts (header-keyed)."""
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".txt"):
        return _parse_csv(content)
    if name.endswith(".xlsx"):
        return _parse_xlsx(content)
    raise ImportError_("unsupported file type (use .csv or .xlsx)")


def _parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig", errors="replace")
    # Sniff delimiter (comma or semicolon — both common in EU exports).
    sample = text[:2048]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    return [dict(row) for row in reader]


def _parse_xlsx(content: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover
        raise ImportError_("openpyxl is required for .xlsx import") from exc
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: list[dict] = []
    for r in rows[1:]:
        if r is None or all(c is None for c in r):
            continue
        out.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
    return out


def normalize_row(raw: dict) -> dict:
    """Map a header-keyed row to canonical fields with light type coercion."""
    canon: dict = {}
    for key, value in raw.items():
        if key is None:
            continue
        field = _ALIAS_TO_CANONICAL.get(str(key).strip().lower())
        if field and value not in (None, ""):
            canon[field] = value

    # Coerce types defensively.
    if "weight_kg" in canon:
        canon["weight_kg"] = _coerce_weight(canon["weight_kg"])
    if "customer_rate" in canon:
        canon["customer_rate"] = _coerce_float(canon["customer_rate"])
    if "adr_required" in canon:
        canon["adr_required"] = str(canon["adr_required"]).strip().lower() in _TRUTHY
    for cc in ("origin_country", "dest_country"):
        if cc in canon:
            canon[cc] = str(canon[cc]).strip().upper()[:2]
    if "currency" in canon:
        canon["currency"] = str(canon["currency"]).strip().upper()[:3] or "EUR"
    for c in ("origin_city", "dest_city", "vehicle_type", "raw_text", "source_reference"):
        if c in canon and canon[c] is not None:
            canon[c] = str(canon[c]).strip()
    return canon


def _coerce_float(value) -> float | None:
    try:
        return float(str(value).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return None


def _coerce_weight(value) -> float | None:
    s = str(value).strip().lower()
    f = _coerce_float(s.replace("t", "").replace("kg", ""))
    if f is None:
        return None
    # "22t" / "22 t" -> kg; bare small numbers under 100 assumed tonnes.
    if "kg" in s:
        return f
    if "t" in s or f < 100:
        return f * 1000
    return f
